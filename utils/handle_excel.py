#!/usr/bin/python3
"""
@File: handle_excel.py
"""

from collections import namedtuple
import os
import platform

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, colors

from common.dir_config import BACKTESTCASEPATH, CSVFILEPATH, EXCELFILEPATH
import pandas as pd


class HandleExcel(object):
    """
    Encapsulate excel tools; Implement read and write operations
    """

    def __init__(self, filename, sheetname=None):
        """Instantiate file attributes and initialize operation file objects"""
        self.filename = filename
        self.sheetname = sheetname
        self.wb = load_workbook(self.filename)
        self.ws = self.wb[
            self.sheetname] if self.sheetname is not None else self.wb.active    # get first sheet
        self.sheet_head_tuple = tuple(
            self.ws.iter_rows(max_row=self.ws.min_row, values_only=True))[0]
        self.cases_list = []    # Define an object to store tuples
        self.Cases = namedtuple("cases", self.sheet_head_tuple)    # Create a namedtuple class

    def get_all_cases(self):
        """Get test cases of all rows in Excel"""
        for tuple_data in self.ws.iter_rows(min_row=self.ws.min_row + 1, values_only=True):    # Each traversal returns a tuple composed of all cell values in a row
            self.cases_list.append(self.Cases(*tuple_data))
        return self.cases_list

    def get_one_case(self, row):
        """Specify whether the test case that returns a line fails to execute in the future. You can try again by obtaining the caseid"""
        if isinstance(row, int) and (self.ws.min_row + 1 <= row <= self.ws.max_row):
            return tuple(self.ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]
        else:
            self.logger.error("The incoming line number is incorrect. It should be an integer greater than 1...")

    def write_file(self, row, actul_result, result_status):
        """The execution case results are written into excel and saved"""
        if isinstance(row, int) and (self.ws.min_row + 1 <= row <= self.ws.max_row):
            self.ws.cell(row=row, column=self.sheet_head_tuple.index(
                "actual") + 1, value=actul_result)
            self.ws.cell(row=row, column=self.sheet_head_tuple.index(
                "result") + 1, value=result_status)
            self.wb.save(self.filename)
        else:
            self.logger.error("Failed to write file, please make sure the file has this cell...")


class Writexcel(object):
    """write datas to excel"""

    def __init__(self, filename):
        """Initializing file object"""
        self.filename = filename
        # If excel backup exists, delete it
        if os.path.exists(BACKTESTCASEPATH):
            os.remove(BACKTESTCASEPATH)
        # Get the current system operation file
        if platform.system() == "Windows" and os.path.exists(self.filename):
            os.renames(self.filename, BACKTESTCASEPATH)
            
        if platform.system() == "Linux" and os.path.exists(self.filename):
            os.renames(self.filename, BACKTESTCASEPATH)

        self.wb = Workbook()
        self.ws = self.wb.active    # activate sheet
        self.ws.cell(1, 1).value = "caseid"
        self.ws.cell(1, 2).value = "title"
        self.ws.cell(1, 3).value = "desc"
        self.ws.cell(1, 4).value = "method"
        self.ws.cell(1, 5).value = "host"
        self.ws.cell(1, 6).value = "port"
        self.ws.cell(1, 7).value = "uri"
        self.ws.cell(1, 8).value = "params"

    def write(self, row_n, col_n, value):
        """Write data, such as (2,3, "hello"), and write data in the second row and the third column"hello\""""
        ft = Font(color=colors.BLUE, size=12, bold=True)
        # Add font style when the judgment value is wrong
        if value in ['fail', 'error'] or col_n == 12:
            self.ws.cell(row_n, col_n).font = ft
        if value == 'pass':
            ft = Font(color=colors.BLACK)
            self.ws.cell(row_n, col_n).font = ft
        self.ws.cell(row_n, col_n).value = value
    
    def xlsx_to_csv_pd(self):
        """xlsx to csv"""
        data_xls = pd.read_excel(self.filename, index_col=0)
        data_xls.to_csv(CSVFILEPATH, encoding='utf-8_sig')    # encoding,utf-8
    
    def save(self):
        self.wb.save(self.filename)


# INIT OBJ
w = Writexcel(EXCELFILEPATH)