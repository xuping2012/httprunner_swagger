# coding:utf-8

import os
from collections import namedtuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, colors


class HandleExcel(object):
    """
    封装excel写入
    """

    def __init__(self, filename, sheetname=None):
        '''实例化文件属性，初始化操作文件对象'''
        self.filename = filename
        self.sheetname = sheetname
        self.wb = load_workbook(self.filename)
        self.ws = self.wb[
            self.sheetname] if self.sheetname is not None else self.wb.active  # 获取第一个表单
        self.sheet_head_tuple = tuple(
            self.ws.iter_rows(max_row=self.ws.min_row, values_only=True))[0]
        self.cases_list = []  # 定义一个存放元组的对象
        self.Cases = namedtuple("cases", self.sheet_head_tuple)  # 创建一个命名元组类

    def get_all_cases(self):
        '''获取excel所有行的测试用例'''
        for tuple_data in self.ws.iter_rows(min_row=self.ws.min_row + 1, values_only=True):  # 每次遍历，返回由某行所有单元格值组成的一个元组
            self.cases_list.append(self.Cases(*tuple_data))
        return self.cases_list

    def get_one_case(self, row):
        '''指定返回某一行的测试用例，是不是以后哪一条测试用例执行失败，可以通过获取caseid来重试'''
        if isinstance(row, int) and (self.ws.min_row + 1 <= row <= self.ws.max_row):
            return tuple(self.ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]
        else:
            #             self.logger.error("传入行号不正确，应为大于1的整数！")
            print(0)

    def write_file(self, row, actul_result, result_status):
        '''执行用例结果写入excel，并保存'''
        if isinstance(row, int) and (self.ws.min_row + 1 <= row <= self.ws.max_row):
            self.ws.cell(row=row, column=self.sheet_head_tuple.index(
                "actual") + 1, value=actul_result)
            self.ws.cell(row=row, column=self.sheet_head_tuple.index(
                "result") + 1, value=result_status)
            self.wb.save(self.filename)
        else:
            #             self.logger.error("写入文件失败，请确认文件有该单元格！")
            print(1)


class Write_excel(object):
    """修改excel数据"""

    def __init__(self, filename):
        '''初始化文件对象'''
        self.filename = filename
        #         创建xlsx文件,如果不存在,顺便写上头
        if not os.path.exists(self.filename):
            self.wb = Workbook()
            self.ws = self.wb.active  # 激活sheet
            self.ws.cell(1, 1).value = "caseid"
            self.ws.cell(1, 2).value = "title"
            self.ws.cell(1, 3).value = "desc"
            self.ws.cell(1, 4).value = "method"
            self.ws.cell(1, 5).value = "host"
            self.ws.cell(1, 6).value = "port"
            self.ws.cell(1, 7).value = "uri"
            self.ws.cell(1, 8).value = "params"
            self.wb.save(filename)
        else:
            self.wb = Workbook()
            self.ws = self.wb.active

    def write(self, row_n, col_n, value):
        """写入数据，如(2,3，"hello"),第二行第三列写入数据"hello\""""
        ft = Font(color=colors.RED, size=12, bold=True)
        # 判断值为错误时添加字体样式
        if value in ['fail', 'error'] or col_n == 12:
            self.ws.cell(row_n, col_n).font = ft
        if value == 'pass':
            ft = Font(color=colors.GREEN)
            self.ws.cell(row_n, col_n).font = ft
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)
